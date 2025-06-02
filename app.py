import os
import uuid
import json
import qrcode
import pandas as pd
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_mail import Mail, Message
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

app = Flask(__name__)
app.secret_key = 'supersecretkey'

# Flask-Mail configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'golucidblack@gmail.com'         # CHANGE THIS
app.config['MAIL_PASSWORD'] = 'lsot pdef doga rgpf'            # CHANGE THIS
app.config['MAIL_DEFAULT_SENDER'] = 'golucidblack@gmail.com'   # CHANGE THIS

mail = Mail(app)

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

REG_FILE = 'registrations.json'
EMAILS_FILE = 'registered_emails.json'
SUBMISSIONS_FILE = 'submissions.xlsx'

# Load registration info
if os.path.exists(REG_FILE):
    with open(REG_FILE, 'r') as f:
        registrations = json.load(f)
else:
    registrations = {}

# Load registered emails
if os.path.exists(EMAILS_FILE):
    with open(EMAILS_FILE, 'r') as f:
        registered_emails = json.load(f)
else:
    registered_emails = []

@app.route('/', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        email = request.form.get('email')
        event_name = request.form.get('event_name')
        file = request.files.get('file')

        if not email or not file or not event_name:
            flash('Email, Event Name, and Excel file are all required.')
            return redirect(request.url)


        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        unique_id = str(uuid.uuid4())
        registration_url = url_for('register', uid=unique_id, _external=True)

        # Save mapping
        registrations[unique_id] = {
            'excel_path': filepath,
            'uploader_email': email,
            'event_name': event_name,
            'shared_sent': False
        }
        with open(REG_FILE, 'w') as f:
            json.dump(registrations, f)

        # Generate QR code
        qr = qrcode.make(registration_url)
        img_io = BytesIO()
        qr.save(img_io, 'PNG')
        img_io.seek(0)

        msg = Message('Your Registration QR Code', recipients=[email])
        msg.body = f'''
        Scan this QR to register: {registration_url}

        When you're ready to share the final registration list with all participants, visit:
        {url_for('share', uid=unique_id, _external=True)}
        '''
        msg.attach('qrcode.png', 'image/png', img_io.read())
        mail.send(msg)


        flash('QR code sent to your email!')
        return render_template('admin_success.html', share_url=url_for('share', uid=unique_id))

        #return redirect(url_for('home'))

    return render_template('form.html')


@app.route('/register/<uid>', methods=['GET', 'POST'])
def register(uid):
    if uid not in registrations:
        flash("Invalid registration link.")
        return redirect(url_for('home'))

    excel_path = registrations[uid]['excel_path']
    event_name = registrations[uid].get('event_name', 'event')
    safe_event_name = event_name.replace(" ", "_").lower()
    submission_file = f'submissions_{safe_event_name}_{uid}.xlsx'


    if request.method == 'POST':
        email = request.form.get('email')
        name = request.form.get('name')
        interest = request.form.get('interest')
        organization = request.form.get('organization')
        linkedin = request.form.get('linkedin').strip()

        if linkedin and not linkedin.startswith('http'):
            linkedin = 'https://' + linkedin

        # Check email in uploaded Excel
        try:
            df = pd.read_excel(excel_path)
            if email not in df.values:
                flash("Email not found in uploaded Excel file.")
                return redirect(url_for('register', uid=uid))
        except Exception as e:
            flash(f"Error reading uploaded file: {e}")
            return redirect(url_for('register', uid=uid))

        # Save submission to Excel
        new_entry = pd.DataFrame([{
            'Name': name,
            'Interest': interest,
            'Organization': organization,
            'LinkedIn': linkedin
        }])

        if os.path.exists(submission_file):
            wb = load_workbook(submission_file)
            ws = wb.active
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(['Name', 'Interest', 'Organization', 'LinkedIn'])  # header row

# Append new row with hyperlink
        row = [name, interest, organization, linkedin]
        ws.append(row)

# Make last LinkedIn cell a real hyperlink
        last_row = ws.max_row
        link_cell = ws.cell(row=last_row, column=4)
        link_cell.hyperlink = linkedin
        link_cell.style = "Hyperlink"
        link_cell.font = Font(color="0000EE", underline="single")

        wb.save(submission_file)

        # Store email for sharing access later
        if email not in registered_emails:
            registered_emails.append(email)
            with open(EMAILS_FILE, 'w') as f:
                json.dump(registered_emails, f)

        flash("Submitted successfully!")
        return redirect(url_for('register', uid=uid))

    return render_template('register.html', uid=uid)


@app.route('/share/<uid>', methods=['GET','POST'])
def share(uid):
    if uid not in registrations:
        flash("Invalid access.")
        return redirect(url_for('home'))

    if registrations[uid].get('shared_sent'):
        flash("Link has already been shared.")
        return redirect(url_for('home'))
    
    event_name = registrations[uid].get('event_name', 'event')
    safe_event_name = event_name.replace(" ", "_").lower()
    submission_file = f'submissions_{safe_event_name}_{uid}.xlsx'

    if not os.path.exists(submission_file):
        flash("Submission file not found.")
        return redirect(url_for('home'))

    # Email all registered users
    try:
        # Include uploader email
        uploader_email = registrations[uid].get('uploader_email')
        all_recipients = set(registered_emails)
        all_recipients.add(uploader_email)

        for recipient in all_recipients:
            msg = Message('Registration Sheet Access', recipients=[recipient])
            msg.body = f"You have access to the full list of registered users for {event_name}."
            with open(submission_file, 'rb') as f:
                msg.attach(f'{event_name}_registrations.xlsx',
                           'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           f.read())
            mail.send(msg)

        # Mark as shared
        registrations[uid]['shared_sent'] = True
        with open(REG_FILE, 'w') as f:
            json.dump(registrations, f)

        flash("Sheet shared with all registered users!")
    except Exception as e:
        flash(f"Error sending emails: {e}")

    return redirect(url_for('home'))


@app.route('/download')
def download():
    if os.path.exists(SUBMISSIONS_FILE):
        return send_file(SUBMISSIONS_FILE, as_attachment=True)
    else:
        flash("Submission file not found.")
        return redirect(url_for('home'))

if __name__ == '__main__':
    app.run(debug=True)
